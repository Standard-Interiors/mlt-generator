import shutil
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from models.estimate import ParsedEstimate
from models.finish_schedule import CrossRefResult
import config


# Column mapping for "Sample & Material mngmt log" sheet, row 15+ data rows
# These are 1-indexed column numbers matching the template
COL_LOCATION = 2        # B
COL_FINISH_SCHED = 3    # C
COL_VENDOR = 4          # D
COL_SELECTION = 5       # E
COL_COLOR = 6           # F
COL_SIZE = 7            # G
COL_THICKNESS = 8       # H
COL_GROUT_COLOR = 9     # I
COL_GROUT_JOINT = 10    # J
COL_ADHESIVE = 11       # K
COL_INSTALL_TYPE = 12   # L
COL_INSTALL_PATTERN = 13  # M
COL_BOX_QTY = 14        # N
COL_BID_PRICE = 25      # Y
COL_VERIFIED_QTY = 27   # AA
COL_WASTE_PCT = 28      # AB
COL_UOM = 33            # AG
COL_NOTES = 43          # AQ

DATA_START_ROW = 16     # First data row (after header row 15)
PROJECT_NAME_CELL = "C14"
START_DATE_CELL = "E14"
SHEET_NAME = "Sample & Material mngmt log"

# Highlight fills
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD_FONT = Font(bold=True)


def _default_waste_pct(product_code: str) -> float:
    """Return a default waste percentage based on material type."""
    code = product_code.upper()
    if code.startswith("CPT") or code.startswith("C-"):
        return 0.05  # 5% for carpet
    elif code.startswith("TL") or code.startswith("T-") or code.startswith("TB-"):
        return 0.10  # 10% for tile
    elif code.startswith("CWT") or code.startswith("SWT"):
        return 0.10  # 10% for wall tile / shower tile
    else:
        return 0.05  # 5% default


def fill_mlt(
    estimate: ParsedEstimate,
    start_date: str,
    project_name: str = "",
    cross_ref: CrossRefResult = None,
) -> str:
    """Fill the MLT template with parsed estimate data and optional cross-reference results.

    Args:
        estimate: Parsed estimate data
        start_date: Project start date as string (YYYY-MM-DD)
        project_name: Optional override for project name
        cross_ref: Optional cross-reference results for discrepancy highlighting

    Returns:
        Path to the generated MLT file
    """
    name = project_name or estimate.project.project_name or "Project"
    safe_name = "".join(c for c in name if c.isalnum() or c in " -_").strip()
    output_filename = f"{safe_name} - MLT.xlsx"
    output_path = os.path.join(config.OUTPUT_FOLDER, output_filename)

    # Copy template to output
    shutil.copy(config.TEMPLATE_PATH, output_path)

    # Open and fill
    wb = openpyxl.load_workbook(output_path)
    ws = wb[SHEET_NAME]

    # Fill project info
    ws[PROJECT_NAME_CELL] = name

    # Parse and set start date
    try:
        dt = datetime.strptime(start_date, "%Y-%m-%d")
        ws[START_DATE_CELL] = dt
    except (ValueError, TypeError):
        ws[START_DATE_CELL] = start_date

    # Build discrepancy lookup for highlighting
    disc_map = {}  # code -> Discrepancy
    plans_only_items = []
    if cross_ref:
        for d in cross_ref.discrepancies:
            disc_map[d.finish_code.upper()] = d
        plans_only_items = cross_ref.plans_only

    # Fill material rows
    current_row = DATA_START_ROW
    for mat in estimate.materials:
        _write_material_row(ws, current_row, mat)
        current_row += 1

        # Check if this code has a discrepancy - add plans data row
        code_upper = mat.product_code.upper()
        if code_upper in disc_map:
            disc = disc_map[code_upper]
            _write_discrepancy_row(ws, current_row, disc)
            current_row += 1

    # Add plans-only items at the end
    if plans_only_items:
        # Add a blank separator row
        current_row += 1
        for disc in plans_only_items:
            _write_plans_only_row(ws, current_row, disc)
            current_row += 1

    wb.save(output_path)
    wb.close()

    return output_path


def _write_material_row(ws, row: int, mat):
    """Write a single material row to the worksheet."""
    ws.cell(row=row, column=COL_LOCATION, value=mat.location)
    ws.cell(row=row, column=COL_FINISH_SCHED, value=mat.product_code)
    ws.cell(row=row, column=COL_VENDOR, value=mat.vendor)
    ws.cell(row=row, column=COL_SELECTION, value=mat.selection)
    ws.cell(row=row, column=COL_COLOR, value=mat.color)
    ws.cell(row=row, column=COL_SIZE, value=mat.size)
    ws.cell(row=row, column=COL_THICKNESS, value=mat.thickness)
    ws.cell(row=row, column=COL_GROUT_COLOR, value=mat.grout_color)
    ws.cell(row=row, column=COL_GROUT_JOINT, value=mat.grout_joint_size)
    ws.cell(row=row, column=COL_ADHESIVE, value=mat.adhesive)
    ws.cell(row=row, column=COL_INSTALL_TYPE, value=mat.install_type)
    ws.cell(row=row, column=COL_INSTALL_PATTERN, value=mat.install_pattern)
    if mat.box_qty:
        ws.cell(row=row, column=COL_BOX_QTY, value=mat.box_qty)

    # Additional columns
    if mat.dollar_amount:
        ws.cell(row=row, column=COL_BID_PRICE, value=mat.dollar_amount)
    if mat.quantity:
        ws.cell(row=row, column=COL_VERIFIED_QTY, value=mat.quantity)
    ws.cell(row=row, column=COL_WASTE_PCT, value=_default_waste_pct(mat.product_code))
    if mat.unit:
        ws.cell(row=row, column=COL_UOM, value=mat.unit)
    if mat.notes:
        ws.cell(row=row, column=COL_NOTES, value=mat.notes)


def _write_discrepancy_row(ws, row: int, disc):
    """Write a highlighted discrepancy row showing plans data."""
    ws.cell(row=row, column=COL_FINISH_SCHED, value=disc.finish_code)
    ws.cell(row=row, column=COL_VENDOR, value=disc.plans_manufacturer)
    ws.cell(row=row, column=COL_SELECTION, value=disc.plans_product)
    ws.cell(row=row, column=COL_COLOR, value=disc.plans_color)
    ws.cell(row=row, column=COL_SIZE, value=disc.plans_dimensions)

    rooms_str = ", ".join(disc.rooms[:5]) if disc.rooms else ""
    note = f"PLANS DATA - REVIEW: {disc.notes}"
    if rooms_str:
        note += f" | Rooms: {rooms_str}"
    ws.cell(row=row, column=COL_NOTES, value=note)

    # Highlight entire row yellow
    for col in range(1, 44):
        cell = ws.cell(row=row, column=col)
        cell.fill = YELLOW_FILL


def _write_plans_only_row(ws, row: int, disc):
    """Write a highlighted row for items in plans but not in estimate."""
    rooms_str = ", ".join(disc.rooms[:5]) if disc.rooms else ""
    ws.cell(row=row, column=COL_LOCATION, value=rooms_str or "See Plans")
    ws.cell(row=row, column=COL_FINISH_SCHED, value=disc.finish_code)
    ws.cell(row=row, column=COL_VENDOR, value=disc.plans_manufacturer)
    ws.cell(row=row, column=COL_SELECTION, value=disc.plans_product)
    ws.cell(row=row, column=COL_COLOR, value=disc.plans_color)
    ws.cell(row=row, column=COL_SIZE, value=disc.plans_dimensions)

    note = f"IN PLANS BUT NOT IN ESTIMATE - REVIEW"
    if rooms_str:
        note += f" | Rooms: {rooms_str}"
    ws.cell(row=row, column=COL_NOTES, value=note)

    # Highlight entire row red
    for col in range(1, 44):
        cell = ws.cell(row=row, column=col)
        cell.fill = RED_FILL
        cell.font = BOLD_FONT
