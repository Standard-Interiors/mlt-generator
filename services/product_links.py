import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models.estimate import ParsedEstimate
from services.web_lookup import lookup_product_url
import config


# Section ordering for the Product Data Links sheet
SECTION_ORDER = [
    "UNIT MATERIALS",
    "COMMON AREA CARPET & CARPET TILE",
    "COMMON RESILIENT & WOOD",
    "COMMON LVT",
    "COMMON AREA FLOOR & WALL TILE",
    "OTHER",
]

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_product_links(estimate: ParsedEstimate, project_name: str = "") -> str:
    """Generate a Product Data Links spreadsheet.

    Args:
        estimate: Parsed estimate data
        project_name: Optional project name override

    Returns:
        Path to the generated file
    """
    name = project_name or estimate.project.project_name or "Project"
    safe_name = "".join(c for c in name if c.isalnum() or c in " -_").strip()
    output_filename = f"{safe_name} - Product Data Links.xlsx"
    output_path = os.path.join(config.OUTPUT_FOLDER, output_filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Data Links"

    # Title row
    ws["A1"] = f"{name} - Product Data Sheet Links"
    ws["A1"].font = Font(bold=True, size=14)

    # Subtitle with customer/address/quote info
    subtitle_parts = []
    if estimate.project.customer_name:
        subtitle_parts.append(estimate.project.customer_name)
    if estimate.project.address:
        subtitle_parts.append(estimate.project.address)
    if estimate.project.quote_number:
        subtitle_parts.append(f"Quote #{estimate.project.quote_number}")
    ws["A2"] = " | ".join(subtitle_parts) if subtitle_parts else ""

    # Header row
    headers = ["#", "Finish Code", "Manufacturer", "Product", "Color/Style", "Product Data Link"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Group materials by section
    sections = {}
    for mat in estimate.materials:
        section = mat.section or "OTHER"
        if section not in sections:
            sections[section] = []
        sections[section].append(mat)

    # Write materials grouped by section
    current_row = 5
    item_num = 1

    for section_name in SECTION_ORDER:
        if section_name not in sections:
            continue

        # Section header row
        ws.cell(row=current_row, column=1, value=section_name)
        ws.cell(row=current_row, column=1).font = SECTION_FONT
        ws.cell(row=current_row, column=1).fill = SECTION_FILL
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).fill = SECTION_FILL
        current_row += 1

        for mat in sections[section_name]:
            # Look up product URL
            url = lookup_product_url(mat.vendor, mat.selection, mat.color)

            ws.cell(row=current_row, column=1, value=item_num)
            ws.cell(row=current_row, column=2, value=mat.product_code)
            ws.cell(row=current_row, column=3, value=mat.vendor)
            ws.cell(row=current_row, column=4, value=mat.selection)
            ws.cell(row=current_row, column=5, value=mat.color)
            ws.cell(row=current_row, column=6, value=url or "ENTER URL")

            # Apply borders
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = THIN_BORDER

            # Make URL a hyperlink if it's a real URL
            if url and url.startswith("http"):
                ws.cell(row=current_row, column=6).hyperlink = url
                ws.cell(row=current_row, column=6).font = Font(color="0563C1", underline="single")

            current_row += 1
            item_num += 1

    # Auto-fit column widths (approximate)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 60

    wb.save(output_path)
    wb.close()

    return output_path
